import json
import logging
import os
import uuid
from decimal import Decimal, InvalidOperation
from django.http import JsonResponse, HttpResponseNotAllowed, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.db import transaction

from .models import Project, Attachment, SupplierAccess, SupplierAccessRound, SupplierInteractionFile, Quote, QuoteLine

logger = logging.getLogger(__name__)


def _normalize_name(s):
    """Collapse whitespace and strip for robust supplier name matching."""
    return ' '.join(str(s).split()).strip()


# Backward-compatible aliases during refactor (moved to rfq/api_common.py)
from .api_common import (
    require_buyer_auth as _require_buyer_auth,
    require_same_origin_for_unsafe as _require_same_origin_for_unsafe,
    get_buyer_username as _get_buyer_username,
    json_body as _json_body,
)


def health(request):
    return JsonResponse({'ok': True})


@csrf_exempt
def projects_collection(request):
    auth_err = _require_buyer_auth(request)
    if auth_err:
        return auth_err

    csrf_err = _require_same_origin_for_unsafe(request)
    if csrf_err:
        return csrf_err

    if request.method == 'GET':
        projects = [p.as_dict() for p in Project.objects.all()]
        return JsonResponse({'projects': projects})

    if request.method == 'POST':
        payload = _json_body(request)
        if not isinstance(payload, dict):
            return JsonResponse({'error': 'Invalid JSON body'}, status=400)

        proj = payload
        pid = str(proj.get('id') or uuid.uuid4().hex)
        name = str(proj.get('name') or 'Untitled')[:255]
        with transaction.atomic():
            obj, _created = Project.objects.select_for_update().get_or_create(id=pid, defaults={'name': name, 'data': proj})
            obj.name = name
            obj.data = proj
            obj.save()
        return JsonResponse({'project': obj.as_dict()})

    return HttpResponseNotAllowed(['GET', 'POST'])


@csrf_exempt
def project_detail(request, project_id: str):
    auth_err = _require_buyer_auth(request)
    if auth_err:
        return auth_err

    csrf_err = _require_same_origin_for_unsafe(request)
    if csrf_err:
        return csrf_err

    pid = str(project_id)

    if request.method == 'GET':
        try:
            obj = Project.objects.get(id=pid)
        except Project.DoesNotExist:
            return JsonResponse({'error': 'Not found'}, status=404)
        return JsonResponse({'project': obj.as_dict()})

    if request.method in ('PUT', 'PATCH'):
        payload = _json_body(request)
        if not isinstance(payload, dict):
            return JsonResponse({'error': 'Invalid JSON body'}, status=400)
        proj = payload
        name = str(proj.get('name') or 'Untitled')[:255]
        with transaction.atomic():
            obj, _created = Project.objects.select_for_update().get_or_create(id=pid, defaults={'name': name, 'data': proj})
            obj.name = name
            obj.data = proj
            obj.save()
        return JsonResponse({'project': obj.as_dict()})

    if request.method == 'DELETE':
        Project.objects.filter(id=pid).delete()
        return JsonResponse({'ok': True})

    return HttpResponseNotAllowed(['GET', 'PUT', 'PATCH', 'DELETE'])


@csrf_exempt
def projects_bulk(request):
    auth_err = _require_buyer_auth(request)
    if auth_err:
        return auth_err

    csrf_err = _require_same_origin_for_unsafe(request)
    if csrf_err:
        return csrf_err

    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])

    payload = _json_body(request)
    projects = None
    if isinstance(payload, dict) and isinstance(payload.get('projects'), list):
        projects = payload.get('projects')
    elif isinstance(payload, list):
        projects = payload

    if not isinstance(projects, list):
        body_len = len(request.body) if request.body else 0
        logger.warning("[RFQ API] projects_bulk 400: payload type=%s, body_len=%s", type(payload).__name__, body_len)
        return JsonResponse({'error': f'Invalid payload (expected {{projects:[...]}}), got {type(payload).__name__}, body_len={body_len}'}, status=400)

    upserted = 0
    deleted = 0
    incoming_ids = set()

    with transaction.atomic():
        # Upsert all incoming projects
        for proj in projects:
            if not isinstance(proj, dict):
                continue
            pid = str(proj.get('id') or uuid.uuid4().hex)
            incoming_ids.add(pid)
            name = str(proj.get('name') or 'Untitled')[:255]
            obj, _created = Project.objects.select_for_update().get_or_create(id=pid, defaults={'name': name, 'data': proj})
            obj.name = name
            obj.data = proj
            obj.save()
            upserted += 1

        # Delete projects not in incoming list (they were deleted locally)
        existing_ids = set(Project.objects.values_list('id', flat=True))
        ids_to_delete = existing_ids - incoming_ids
        if ids_to_delete:
            deleted = Project.objects.filter(id__in=ids_to_delete).delete()[0]

    return JsonResponse({'ok': True, 'upserted': upserted, 'deleted': deleted})


@csrf_exempt
def projects_reset(request):
    auth_err = _require_buyer_auth(request)
    if auth_err:
        return auth_err

    csrf_err = _require_same_origin_for_unsafe(request)
    if csrf_err:
        return csrf_err

    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])
    Project.objects.all().delete()
    return JsonResponse({'ok': True})


@csrf_exempt
def project_attachments(request, project_id: str):
    """List / upload attachments for a project."""
    auth_err = _require_buyer_auth(request)
    if auth_err:
        return auth_err

    csrf_err = _require_same_origin_for_unsafe(request)
    if csrf_err:
        return csrf_err

    try:
        proj = Project.objects.get(id=str(project_id))
    except Project.DoesNotExist:
        return JsonResponse({'error': 'Project not found'}, status=404)

    if request.method == 'GET':
        atts = [a.as_dict() for a in Attachment.objects.filter(project=proj).order_by('-uploaded_at')]
        return JsonResponse({'attachments': atts})

    if request.method == 'POST':
        f = request.FILES.get('file')
        if not f:
            return JsonResponse({'error': 'Missing file'}, status=400)
        kind = str(request.POST.get('kind') or '').strip()[:32]
        att = Attachment(id=uuid.uuid4().hex, project=proj, file=f, kind=kind)
        att.save()
        return JsonResponse({'attachment': att.as_dict()})

    return HttpResponseNotAllowed(['GET', 'POST'])


@csrf_exempt
def project_attachment_detail(request, project_id: str, attachment_id: str):
    """Delete a single attachment."""
    auth_err = _require_buyer_auth(request)
    if auth_err:
        return auth_err

    csrf_err = _require_same_origin_for_unsafe(request)
    if csrf_err:
        return csrf_err

    try:
        att = Attachment.objects.get(id=str(attachment_id), project_id=str(project_id))
    except Attachment.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)

    if request.method == 'DELETE':
        # delete file from storage as well
        try:
            if att.file:
                att.file.delete(save=False)
        except Exception:
            pass
        att.delete()
        return JsonResponse({'ok': True})

    return HttpResponseNotAllowed(['DELETE'])


@csrf_exempt
def export_data(request):
    auth_err = _require_buyer_auth(request)
    if auth_err:
        return auth_err

    csrf_err = _require_same_origin_for_unsafe(request)
    if csrf_err:
        return csrf_err

    """Generate XLSX/PDF/CSV export based on selected options.

    Body (JSON):
      - project_ids: [..]
      - format: xlsx|pdf|csv
      - include_items, include_item_suppliers, include_price_breaks, include_rfqs, include_attachments
      - suppliers_mode: all|main
    """
    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])

    payload = _json_body(request)
    if not isinstance(payload, dict):
        return JsonResponse({'error': 'Invalid JSON body'}, status=400)

    project_ids = payload.get('project_ids') or []
    if not isinstance(project_ids, list) or not project_ids:
        return JsonResponse({'error': 'project_ids is required'}, status=400)

    fmt = str(payload.get('format') or 'xlsx').strip().lower()
    if fmt not in ('xlsx', 'pdf', 'csv'):
        fmt = 'xlsx'

    include_items = bool(payload.get('include_items', True))
    include_item_suppliers = bool(payload.get('include_item_suppliers', True))
    include_price_breaks = bool(payload.get('include_price_breaks', True))
    include_rfqs = bool(payload.get('include_rfqs', True))
    include_attachments = bool(payload.get('include_attachments', False))
    suppliers_mode = str(payload.get('suppliers_mode') or 'all').strip().lower()
    if suppliers_mode not in ('all', 'main'):
        suppliers_mode = 'all'

    # Fetch projects
    projects = list(Project.objects.filter(id__in=[str(x) for x in project_ids]))
    # keep order of requested ids
    proj_by_id = {str(p.id): p for p in projects}
    projects = [proj_by_id.get(str(pid)) for pid in project_ids if proj_by_id.get(str(pid))]
    if not projects:
        return JsonResponse({'error': 'No projects found for given IDs'}, status=404)

    def safe_num(v):
        try:
            s = str(v).replace(',', '.').strip()
            if not s:
                return 0.0
            return float(s)
        except Exception:
            return 0.0

    def as_date(v):
        s = str(v or '').strip()
        return s.split('T')[0] if s else ''

    MAX_TIERS = 10  # standard export tiers (Qty/Price 1..10)

    def _qty_cols(it):
        cols = {}
        for i in range(1, MAX_TIERS + 1):
            cols[f'qty_{i}'] = it.get(f'qty_{i}') or ''
        extra = []
        for k, v in (it or {}).items():
            if not isinstance(k, str) or not k.startswith('qty_'):
                continue
            try:
                idx = int(k.split('_')[1])
            except Exception:
                continue
            if idx > MAX_TIERS:
                sv = str(v).strip()
                if sv:
                    extra.append((idx, sv))
        extra.sort(key=lambda t: t[0])
        cols['qty_next'] = " | ".join([f"qty_{i}={v}" for i, v in extra])
        return cols

    def _pick_main_supplier(it, main_name):
        sups = it.get('suppliers') or []
        if not isinstance(sups, list):
            return None
        for s in sups:
            if isinstance(s, dict) and (s.get('isMain') or s.get('is_main')):
                return s
        if main_name:
            mn = str(main_name).strip().lower()
            for s in sups:
                if not isinstance(s, dict):
                    continue
                sname = (s.get('name') or s.get('supplier_name') or s.get('supplier') or '')
                if str(sname).strip().lower() == mn:
                    return s
        for s in sups:
            if isinstance(s, dict):
                return s
        return None

    def _supplier_price_list(s):
        if not isinstance(s, dict):
            return []
        prices = s.get('prices')
        if isinstance(prices, list) and prices:
            tmp = []
            for x in prices:
                if not isinstance(x, dict):
                    continue
                q = safe_num(x.get('qty') or x.get('quantity') or x.get('break') or x.get('qty_break') or '')
                pv = x.get('price') or x.get('unit_price') or x.get('unit') or x.get('value') or ''
                tmp.append((q, pv))
            tmp.sort(key=lambda t: (t[0] == 0, t[0]))
            return [pv for _, pv in tmp]
        out = []
        out.append(s.get('price_1') or s.get('price') or '')
        for i in range(2, 31):
            out.append(s.get(f'price_{i}') or '')
        while out and str(out[-1]).strip() == '':
            out.pop()
        return out

    def _price_cols(it, s):
        cols = {}
        plist = _supplier_price_list(s)
        for i in range(1, MAX_TIERS + 1):
            v = ''
            if i - 1 < len(plist):
                v = plist[i - 1]
            if not str(v).strip():
                v = (it.get('price_1') or it.get('price') or '') if i == 1 else (it.get(f'price_{i}') or '')
            cols[f'price_{i}'] = v or ''
        extra = []
        for idx in range(MAX_TIERS + 1, len(plist) + 1):
            v = plist[idx - 1]
            if str(v).strip():
                extra.append((idx, str(v).strip()))
        cols['price_next'] = " | ".join([f"price_{i}={v}" for i, v in extra])
        return cols

    # Flatten data
    rows_projects = []
    rows_items = []
    rows_item_sups = []
    rows_suppliers = []
    rows_rfqs = []
    rows_atts = []

    supplier_agg = {}  # (proj_id, sup_name) -> dict

    for p in projects:
        pdata = p.data or {}
        rows_projects.append({
            'project_id': p.id,
            'project_name': p.name,
            'project_status': pdata.get('project_status') or pdata.get('status') or '',
            'created_at': as_date(pdata.get('created_at') or p.created_at),
            'updated_at': as_date(pdata.get('updated_at') or p.updated_at),
            'deadline': (pdata.get('dates') or {}).get('deadline') or pdata.get('deadline') or '',
            'sent_to': pdata.get('sent_to') or '',
            'notes_count': len(pdata.get('notes') or []),
            'items_count': len(pdata.get('items') or []),
            'rfq_bundles': len(pdata.get('rfqBatches') or pdata.get('rfq_batches') or []),
        })

        items = pdata.get('items') or []
        batches = pdata.get('rfqBatches') or pdata.get('rfq_batches') or []

        if include_rfqs:
            for b in batches:
                if not isinstance(b, dict): 
                    continue
                rows_rfqs.append({
                    'project_id': p.id,
                    'project_name': p.name,
                    'bundle_id': b.get('id') or '',
                    'supplier': b.get('supplier_name') or b.get('supplier') or '',
                    'status': b.get('status') or '',
                    'created_at': as_date(b.get('created_at')),
                    'due_date': b.get('due_date') or '',
                    'items_count': len(b.get('items') or []),
                    'currency': b.get('currency') or '',
                    'note': b.get('note') or '',
                })

        if include_items:
            for it in items:
                if not isinstance(it, dict): 
                    continue
                dn = it.get('item_drawing_no') or it.get('drawing_no') or it.get('line') or ''
                main_sup = it.get('supplier') or ''
                main_price = it.get('price_1') or it.get('price') or ''
                main_cur = it.get('currency') or 'EUR'
                rows_items.append({
                    'project_id': p.id,
                    'project_name': p.name,
                    'drawing_no': dn,
                    'description': it.get('description') or '',
                    'manufacturer': it.get('manufacturer') or '',
                    'mpn': it.get('mpn') or it.get('MPN') or it.get('manufacturer_part_no') or it.get('mfr_part_no') or it.get('part_no') or '',
                    'status': it.get('status') or '',
                    'main_supplier': main_sup,
                    'currency': main_cur,
                    'suppliers_count': len(it.get('suppliers') or []) if isinstance(it.get('suppliers'), list) else 0,
                    'lead_time': it.get('lead_time') or '',
                    'shipping_cost': it.get('shipping_cost') or '',
                })
                # Expand Items sheet with Qty/Price tiers (1..10 + next)
                try:
                    rows_items[-1].update(_qty_cols(it))
                    rows_items[-1].update(_price_cols(it, _pick_main_supplier(it, main_sup)))
                except Exception:
                    pass

                if include_item_suppliers:
                    sups = it.get('suppliers') or []
                    if not isinstance(sups, list):
                        sups = []
                    for s in sups:
                        if not isinstance(s, dict):
                            continue
                        sname = s.get('name') or s.get('supplier_name') or s.get('supplier') or ''
                        if not sname:
                            continue
                        is_main = bool(s.get('isMain') or (main_sup and str(main_sup).strip().lower() == str(sname).strip().lower()))
                        if suppliers_mode == 'main' and not is_main:
                            continue

                        row = {
                            'project_id': p.id,
                            'project_name': p.name,
                            'drawing_no': dn,
                            'supplier': sname,
                            'is_main': 'YES' if is_main else '',
                            'status': s.get('status') or '',
                            'currency': s.get('currency') or main_cur,
                            'price_raw': s.get('price') or '',
                            'moq': s.get('moq') or '',
                            'mov': s.get('mov') or '',
                            'lead_time': s.get('lead_time') or '',
                            'shipping_cost': s.get('shipping_cost') or s.get('shipping') or '',
                            'payment_terms': s.get('payment_terms') or '',
                            'incoterms': s.get('incoterms') or '',
                            'valid_until': s.get('valid_until') or s.get('quote_valid_until') or '',
                            'notes': s.get('notes') or '',
                        }
                        # Standardize Qty/Price tiers in export (1..10 + next)
                        try:
                            row.update(_qty_cols(it))
                            if include_price_breaks:
                                row.update(_price_cols(it, s))
                            else:
                                base_prices = {f'price_{i}': '' for i in range(1, MAX_TIERS + 1)}
                                base_prices['price_1'] = s.get('price_1') or s.get('price') or ''
                                base_prices['price_next'] = ''
                                row.update(base_prices)
                        except Exception:
                            pass
                        rows_item_sups.append(row)


                        # aggregate per supplier
                        key = (str(p.id), str(sname).strip())
                        agg = supplier_agg.get(key)
                        if not agg:
                            agg = supplier_agg[key] = {
                                'project_id': p.id,
                                'project_name': p.name,
                                'supplier': str(sname).strip(),
                                'items': set(),
                                'main_items': set(),
                                'quoted_items': set(),
                                'avg_price_eur_sum': 0.0,
                                'avg_price_eur_cnt': 0,
                            }
                        if dn:
                            agg['items'].add(str(dn))
                            if is_main:
                                agg['main_items'].add(str(dn))
                            price_val = safe_num(s.get('price_1') or s.get('price') or 0)
                            if price_val > 0:
                                agg['quoted_items'].add(str(dn))
                                agg['avg_price_eur_sum'] += price_val
                                agg['avg_price_eur_cnt'] += 1

        if include_attachments:
            for a in Attachment.objects.filter(project=p).order_by('-uploaded_at'):
                rows_atts.append({
                    'project_id': p.id,
                    'project_name': p.name,
                    'attachment_id': a.id,
                    'kind': a.kind,
                    'file': a.file.name if a.file else '',
                    'uploaded_at': as_date(a.uploaded_at),
                })

    # finalize supplier agg
    for agg in supplier_agg.values():
        rows_suppliers.append({
            'project_id': agg['project_id'],
            'project_name': agg['project_name'],
            'supplier': agg['supplier'],
            'items_count': len(agg['items']),
            'main_items': len(agg['main_items']),
            'quoted_items': len(agg['quoted_items']),
            'avg_price': (agg['avg_price_eur_sum'] / agg['avg_price_eur_cnt']) if agg['avg_price_eur_cnt'] else '',
        })


    # --- STANDARDIZE EXPORT HEADERS (ensure tier columns exist on every row) ---
    try:
        _tier_qty_keys = [f'qty_{i}' for i in range(1, MAX_TIERS + 1)] + ['qty_next']
        _tier_price_keys = [f'price_{i}' for i in range(1, MAX_TIERS + 1)] + ['price_next']
        for _r in rows_items:
            for _k in _tier_qty_keys + _tier_price_keys:
                _r.setdefault(_k, '')
        for _r in rows_item_sups:
            for _k in _tier_qty_keys + _tier_price_keys:
                _r.setdefault(_k, '')
    except Exception:
        pass

    # CSV: keep it simple (items only)
    if fmt == 'csv':
        import csv
        import io
        out = io.StringIO()
        if not include_items:
            out.write("Nothing selected for export\n")
        else:
            headers = (
                ['project_id','project_name','drawing_no','description','manufacturer','mpn','status',
                 'main_supplier','currency','suppliers_count','lead_time','shipping_cost']
                + [f'qty_{i}' for i in range(1, MAX_TIERS + 1)] + ['qty_next']
                + [f'price_{i}' for i in range(1, MAX_TIERS + 1)] + ['price_next']
            )
            w = csv.DictWriter(out, fieldnames=headers)
            w.writeheader()
            for r in rows_items:
                w.writerow({k: r.get(k, '') for k in headers})
        data = out.getvalue().encode('utf-8-sig')
        resp = HttpResponse(data, content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = 'attachment; filename="rfq_export.csv"'
        return resp

    if fmt == 'pdf':
        # PDF summary (ReportLab)
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas
            from reportlab.lib.units import mm
        except Exception as e:
            return JsonResponse({'error': f'ReportLab not available: {e}'}, status=500)

        import io
        buf = io.BytesIO()
        c = canvas.Canvas(buf, pagesize=A4)
        width, height = A4
        y = height - 18*mm

        c.setFont("Helvetica-Bold", 14)
        c.drawString(18*mm, y, "RFQ Export Summary")
        y -= 10*mm

        c.setFont("Helvetica", 10)
        for rp in rows_projects:
            line = f"{rp['project_name']} ({rp['project_id']}) — Status: {rp.get('project_status','')} — Items: {rp.get('items_count','')} — Bundles: {rp.get('rfq_bundles','')}"
            c.drawString(18*mm, y, line[:120])
            y -= 6*mm
            if y < 20*mm:
                c.showPage()
                y = height - 18*mm
                c.setFont("Helvetica", 10)

        y -= 4*mm
        c.setFont("Helvetica-Bold", 12)
        c.drawString(18*mm, y, "Top suppliers (by quoted items)")
        y -= 8*mm
        c.setFont("Helvetica", 10)
        top = sorted(rows_suppliers, key=lambda x: (x.get('quoted_items') or 0, x.get('items_count') or 0), reverse=True)[:20]
        for s in top:
            line = f"{s['project_name']} — {s['supplier']} — Items: {s['items_count']} — Quoted: {s['quoted_items']} — Main: {s['main_items']}"
            c.drawString(18*mm, y, line[:120])
            y -= 6*mm
            if y < 20*mm:
                c.showPage()
                y = height - 18*mm
                c.setFont("Helvetica", 10)

        c.save()
        buf.seek(0)
        resp = HttpResponse(buf.read(), content_type='application/pdf')
        resp['Content-Disposition'] = 'attachment; filename="rfq_export.pdf"'
        return resp

    # XLSX
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except Exception as e:
        return JsonResponse({'error': f'openpyxl not available: {e}'}, status=500)

    wb = Workbook()
    wb.remove(wb.active)

    def add_sheet(title, rows, headers=None):
        ws = wb.create_sheet(title=title[:31])
        if not rows:
            ws.append(["(no data)"])
            return
        if headers is None:
            # union of keys (deterministic, avoids missing tier columns)
            keyset = set()
            for r in rows:
                if isinstance(r, dict):
                    keyset.update(r.keys())
            headers = list(keyset)
            headers.sort(key=lambda x: str(x))
        ws.append(headers)
        for r in rows:
            ws.append([r.get(h, '') for h in headers])
        # basic column widths
        for i, h in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(i)].width = min(42, max(10, len(str(h)) + 2))

    # Header presets (keep tier columns always visible)
    ITEM_HEADERS_BASE = [
        'project_id','project_name','drawing_no','description','manufacturer','mpn','status',
        'main_supplier','currency','suppliers_count','lead_time','shipping_cost'
    ]
    ITEM_SUP_HEADERS_BASE = [
        'project_id','project_name','drawing_no','supplier','is_main','status','currency',
        'price_raw','moq','mov','lead_time','shipping_cost','payment_terms','incoterms','valid_until','notes'
    ]
    TIER_Q_HEADERS = [f'qty_{i}' for i in range(1, MAX_TIERS + 1)] + ['qty_next']
    TIER_P_HEADERS = [f'price_{i}' for i in range(1, MAX_TIERS + 1)] + ['price_next']
    ITEM_HEADERS = ITEM_HEADERS_BASE + TIER_Q_HEADERS + TIER_P_HEADERS
    ITEM_SUP_HEADERS = ITEM_SUP_HEADERS_BASE + TIER_Q_HEADERS + TIER_P_HEADERS

    add_sheet("Projects", rows_projects)
    if include_items:
        add_sheet("Items", rows_items, headers=ITEM_HEADERS)
    if include_item_suppliers:
        add_sheet("ItemSuppliers", rows_item_sups, headers=ITEM_SUP_HEADERS)
    if include_rfqs:
        add_sheet("RFQs", rows_rfqs)
    add_sheet("Suppliers", rows_suppliers)
    if include_attachments:
        add_sheet("Attachments", rows_atts)

    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="rfq_export.xlsx"'
    return resp

# ---------------------------------------------------------------------------
# Legacy compatibility aliases
# ---------------------------------------------------------------------------
# Supplier + Quotes endpoints were moved to rfq/api_supplier.py and
# rfq/api_quotes.py. Keep aliases here so any out-of-tree imports of
# rfq.views_api.* continue to work.
from . import api_supplier as _api_supplier
from . import api_quotes as _api_quotes

supplier_access_generate = _api_supplier.supplier_access_generate
supplier_access_viewed = _api_supplier.supplier_access_viewed
supplier_portal_save_draft = _api_supplier.supplier_portal_save_draft
supplier_portal_submit = _api_supplier.supplier_portal_submit
supplier_access_approve = _api_supplier.supplier_access_approve
supplier_access_reject = _api_supplier.supplier_access_reject
project_supplier_access_list = _api_supplier.project_supplier_access_list
supplier_interaction_file_download = _api_supplier.supplier_interaction_file_download
supplier_access_request_reopen = _api_supplier.supplier_access_request_reopen
supplier_access_update_items = _api_supplier.supplier_access_update_items
supplier_access_cancel = _api_supplier.supplier_access_cancel
supplier_access_reopen_buyer = _api_supplier.supplier_access_reopen_buyer
supplier_access_bulk_generate = _api_supplier.supplier_access_bulk_generate

quotes_list = _api_quotes.quotes_list
quotes_detail = _api_quotes.quotes_detail
quotes_create = _api_quotes.quotes_create
quotes_update = _api_quotes.quotes_update
quotes_delete = _api_quotes.quotes_delete
quotes_create_from_item = _api_quotes.quotes_create_from_item
quotes_export_to_item = _api_quotes.quotes_export_to_item
quotes_bulk_import = _api_quotes.quotes_bulk_import
quotes_upsert_from_planner = _api_quotes.quotes_upsert_from_planner
